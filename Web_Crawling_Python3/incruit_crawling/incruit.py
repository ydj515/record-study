#-*- coding:utf-8 -*-
import urllib.request
import bs4
import openpyxl

def get_bsobj(url):
    """
    bs_obj를 리턴
    """
    html = urllib.request.urlopen(url)
    bs_obj = bs4.BeautifulSoup(html, "html.parser") # url에 해당하는 html이 bsObj에 들어감

    return bs_obj

def save_excel(my_dict):
    """
    excel 파일로 저장
    """
    # excel 파일 생성
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    index = ["ㅋ_ㅋ"," ","공고 URL","공고 제목", "회사명", "회사 설립일", "기업 규모(대/중/소)","업종", "회사 주소","모집직종", "모집인원", "경력구분", "학력", "고용형태(정규직/인턴/계약직)", "근무지역", "급여", "직급", "모집 시작일", "모집 마감일", "필요서류"]

    for i in range(1,len(index)):
        sheet.cell(row=1, column=i, value=index[i])
    
    row = 2

    for key, values in my_dict.items():
        
        # put hte key in the first column for each key in the dictionary
        sheet.cell(row=row, column=1, value=key)
        column = 2

        for element in values:

            # put the element in each adjacent column for each element in
            sheet.cell(row=row, column=column, value=element)
            column+=1
        row+=1

    # excel save
    workbook.save(filename="호로록.xlsx")

def main():

    # driver = webdriver.Chrome("./chromedriver.exe")
    url = "http://job.incruit.com/jobdb_list/searchjob.asp"
    query = "?ct=1&ty=1&cd=150&page={}"
    
    all_urls = [] # http://job.incruit.com/jobdb_list/searchjob.asp?ct=1&ty=1&cd=150&page=58
    links= []
    result_dic = {}
    

    for i in range(1,2): # 1 ~ 58 url 모두 담기
        query.format(i)
        uri = url + query.format(i)
        all_urls.append(uri)
        print(uri)

    print("=================")

    for i in range(0,len(all_urls)):
        print(all_urls[i])
        bs_obj = get_bsobj(all_urls[i])
        div = bs_obj.find("div",{"class":"n_job_list_default"})
        
        table = div.find("div",{"class":"n_job_list_table_a"})
        
        a_tags = table.findAll("a",{"class":"links"})

        for j in range(0,len(a_tags)):
            links.append(a_tags[j]["href"])
            
        print(len(links))

    for i in range(0,len(links)):
        print("=================")

        reuslt_value_list = [] # 공고 내용을 담는 list

        job_link = get_bsobj(links[i])
        title = job_link.find("div",{"class":"jobview_top_title"})
        job_title = title.find("strong").text
        company_name = title.find("span").text
        
        reuslt_value_list.append(links[i])

        print(job_title.strip().strip()) # 공고 제목
        reuslt_value_list.append(job_title)
        print(company_name.strip().strip()) # 회사 이름
        reuslt_value_list.append(company_name.strip())

        div_left = job_link.find("div",{"class":"jobpost_sider_cpinfo_table"})
        
        try:
            left_dls = div_left.findAll("dl")
        except:
            left_dls = "None"
            
        try:
            born_date = left_dls[0].find("p").text # 설립일
        except:
            born_date = "None"

        try:
            company_size = left_dls[1].find("a").text # 기업 규모(중소, 대)
        except:
            company_size = "None"

        try:
            company_category = left_dls[2].find("p").text # 업종
        except:
            company_category = "None"

        try:
            company_url = left_dls[3].find("a")["href"]
        except:
            company_url = "None"


        print(born_date.strip())
        reuslt_value_list.append(born_date.strip())
        print(company_size.strip())
        reuslt_value_list.append(company_size.strip())
        print(company_category.strip())
        reuslt_value_list.append(company_category.strip())
        print(company_url.strip())
        reuslt_value_list.append(company_url.strip())

        try:
            div_right = job_link.find("div",{"class":"jobpost_sider_jbinfo"})
            divs = div_right.findAll("div",{"class":"jobpost_sider_jbinfo_inlay"})

            right_dls = divs[0].findAll("dl")
            
            try:
                recruitment_work = right_dls[0].find("div",{"class":"inset_ely_lay"}).text # 모집 직종
            
            except:
                recruitment_work = "None"

            try:
                recruitment_how_many = right_dls[1].find("div",{"class":"inset_ely_lay"}).text # 모집 인원
            except:
                recruitment_how_many = "None"
            
            try:
                right_dls = divs[1].findAll("dl")
            except:
                right_dls = "None"
            
            try:
                recruitment_condition = right_dls[0].find("em",{"class":"pt_txt"}).text # 신입, 경력
            except:
                recruitment_condition = "None"
            
            try:
                recruitment_school = right_dls[1].find("em",{"class":"pt_txt"}).text # 학력
            except:
                recruitment_school = "None"

            try:
                right_dls = divs[2].findAll("dl")
                recruitment_form = right_dls[0].find("em",{"class":"pt_txt"}).text # 고용형태(정규직, 인턴)
            except:
                recruitment_form = "None"
            
            try:
                work_location = right_dls[1].find("div",{"class":"inset_ely_lay"}).text # 근무지역
            except:
                work_location = "None"
            
            try:    
                pay = right_dls[2].find("div",{"class":"inset_ely_lay"}).text # 급여
            except:
                pay = "None"
                
            try:
                position = right_dls[3].find("div",{"class":"inset_ely_lay"}).text # 직급
            except:
                position = "None"

            print(recruitment_work.strip())
            reuslt_value_list.append(recruitment_work.strip())
            print(recruitment_how_many.strip())
            reuslt_value_list.append(recruitment_how_many.strip())
            print(recruitment_condition.strip())
            reuslt_value_list.append(recruitment_condition.strip())
            print(recruitment_school.strip())
            reuslt_value_list.append(recruitment_school.strip())
            print(recruitment_form.strip())
            reuslt_value_list.append(recruitment_form.strip())
            print(work_location.strip())
            reuslt_value_list.append(work_location.strip())
            print(pay.strip())
            reuslt_value_list.append(pay.strip())
            print(position.strip())
            reuslt_value_list.append(position.strip())

        except:
            pass

        try:    
            jobview_receipt_layout1 = job_link.find("div",{"class":"jobview_receipt_layout1"})
        
            jobview_receipt_daybox = jobview_receipt_layout1.find("div",{"class":"jobview_receipt_daybox"})
        except:
            pass

        try:
            dds = jobview_receipt_daybox.findAll("dd")
            start_day = dds[0].find("strong").text
            finish_day = dds[1].find("strong").text
        except:
            start_day = "None"
            finish_day = "None"
        
        print(start_day.strip()) # 모집 시작일
        reuslt_value_list.append(start_day.strip())
        print(finish_day.strip()) # 모집 마감일
        reuslt_value_list.append(finish_day.strip())

        jobview_receipt_layout2 = job_link.find("div",{"class":"jobview_receipt_layout2"})
        try:
            require_documents = jobview_receipt_layout2.find("td").text
        except:
            require_documents = "인크루트 이력서"

        print(require_documents.strip()) # 필요 서류
        reuslt_value_list.append(require_documents.strip())

        print(reuslt_value_list)
        result_dic[links[i]] = reuslt_value_list
    
    print(result_dic)

    print(len(links))

    save_excel(result_dic)

if __name__ == "__main__":
    main()