#-*- coding:utf-8 -*-
import openpyxl

# excel 파일 생성
workbook = openpyxl.Workbook()
sheet = workbook.active

dictt = {"http://job.incruit.com/jobdb_info/jobpost.asp?job=1911150001683" : ['http://job.incruit.com/jobdb_info/jobpost.asp?job=1911150001683', '신입/경력 인재 채용', '세화피앤씨', '1988년 01월 08일', '중소기업', '제조업기타', 'http://www.shehwa.co.kr', '정규직', '경기도 > 용인시 처인구', '회사내규', 'None', '2019.11.15 (금)', '2019.12.15 (일) 23:59', '인크루트 이력서'],
        "http://job.incruit.com/jobdb_info/jobpost.asp?job=1910290000405" : ['http://job.incruit.com/jobdb_info/jobpost.asp?job=1910290000405', '쇼핑몰운영(카페24), 온라인마케팅 신입/경력 정규직 채용', '스카이음향','업데이트 요청', '소기업', '소매', 'http://skyav.co.kr/', '정규직', '서울특별시 > 용산구, 전지역, 동작구 외 23건\t\t\t\t\t\t 더보기', '월급, 협의 후 결정', 'None', 'None', 'None', '이력서에 연락처, 희망연봉 기재제출한 서류는 일체 반환하지 않음이력서, 자기소개서서류전형, 면접전형']
        }

index = ["ㅋ_ㅋ","공고 URL","공고 제목", "회사명", "회사 설립일", "기업 규모(대/중/소)","업종", "모집직종", "모집인원", "경력구분", "학력", "고용형태(정규직/인턴/계약직)", "근무지역", "급여", "직급", "모집 시작일", "모집 마감일", "필요서류"]

for i in range(1,len(index)):
    sheet.cell(row=1, column=i, value=index[i])

row = 2

for key, values in dictt.items():
    
    # put hte key in the first column for each key in the dictionary
    sheet.cell(row=row, column=1, value=key)
    column = 2

    for element in values:

        # put the element in each adjacent column for each element in
        sheet.cell(row=row, column=column, value=element)
        column+=1
    row+=1

# excel save
workbook.save(filename="mymymy.xlsx")